"""
code_survey.py
Clean and recode the Online Shopping Habits survey for SPSS multiple linear regression.
Input : /root/.claude/uploads/e94137eb-3d50-4739-9278-ab949b2eb732/93fc5ac7-Online_Shopping_Habits_of_Young_Adults_in_Belgium_Responses_new.xlsx
Output: /home/user/BabyPluto/coded_survey_data.xlsx  (two sheets: Data + Codebook)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings("ignore")

INPUT  = ("/root/.claude/uploads/e94137eb-3d50-4739-9278-ab949b2eb732/"
          "93fc5ac7-Online_Shopping_Habits_of_Young_Adults_in_Belgium_Responses_new.xlsx")
OUTPUT = "/home/user/BabyPluto/coded_survey_data.xlsx"

# ── 1. Load raw data ──────────────────────────────────────────────────────────
raw = pd.read_excel(INPUT)
print(f"Raw shape: {raw.shape}")

# Collector for unmatched values
unmatched = {}

def track_unmatched(col_name, series, mapping):
    """Return mapped series; record any non-NaN values not in mapping."""
    bad = series[~series.isna() & ~series.isin(mapping.keys())]
    if not bad.empty:
        unmatched[col_name] = bad.unique().tolist()
    return series.map(mapping)

n = len(raw)

# ── 2. AGE (col 1) ───────────────────────────────────────────────────────────
def clean_age(val):
    if pd.isna(val):
        return np.nan
    s = str(val).strip().lower()
    s = s.replace("ans", "").replace("years", "").strip()
    try:
        return int(float(s))
    except Exception:
        return np.nan

age_raw = raw.iloc[:, 1].copy()
age = age_raw.apply(clean_age)
bad_age = age_raw[age.isna() & ~age_raw.isna()]
if not bad_age.empty:
    unmatched["Age"] = bad_age.tolist()

# ── 3. SITUATION (col 2) ─────────────────────────────────────────────────────
situation_map = {
    "Student (non-working)": 1,
    "Student (working)":     2,
    "Employed full-time":    3,
    "Employed part-time":    4,
    "Self-employed":         5,
    "Unemployed":            6,
    "Stay-at-home":          7,
}
sit_raw = raw.iloc[:, 2].astype(str).str.strip()
situation = track_unmatched("Situation", sit_raw, situation_map)

# ── 4. RELATIONSHIP STATUS (col 3) ───────────────────────────────────────────
def normalise_rel(val):
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    low = s.lower()
    if low == "yep still single":
        return "Single"
    if low in ("situationship",
               "hella toxic situationship with my ex gf (guess who this is) ",
               "hella toxic situationship with my ex gf (guess who this is)"):
        return "In a relationship"
    return s

rel_raw = raw.iloc[:, 3].copy()
rel_cleaned = rel_raw.apply(normalise_rel)
rel_map = {
    "Single":            1,
    "In a relationship": 2,
    "Cohabiting":        3,
    "Married":           4,
}
relationship = track_unmatched("RelationshipStatus", rel_cleaned, rel_map)

# ── 5. MONTHLY INCOME (col 6) ────────────────────────────────────────────────
income_map = {
    "€0 – €999":         1,
    "€1000 – €1999":     2,
    "€2000 – €2999":     3,
    "€3000 – €3999":     4,
    "€4000+":            5,
    "Prefer not to say": np.nan,
}
inc_raw = raw.iloc[:, 6].astype(str).str.strip()
# Handle "Prefer not to say" as intentional NaN — don't flag it
inc_known = set(income_map.keys())
inc_bad = inc_raw[~inc_raw.isna() & ~inc_raw.isin(inc_known)]
if not inc_bad.empty:
    unmatched["MonthlyIncome"] = inc_bad.unique().tolist()
monthly_income = inc_raw.map(income_map).astype(float)

# ── 6. DISPOSABLE INCOME (col 7) ─────────────────────────────────────────────
disp_map = {
    "€0 – €499":   1,
    "€500 – €999": 2,
    "€1000 – €1499": 3,
    "€1500+":      4,
    "Not sure":    np.nan,
}
disp_raw = raw.iloc[:, 7].astype(str).str.strip()
disp_known = set(disp_map.keys())
disp_bad = disp_raw[~disp_raw.isna() & ~disp_raw.isin(disp_known)]
if not disp_bad.empty:
    unmatched["DisposableIncome"] = disp_bad.unique().tolist()
disposable = disp_raw.map(disp_map).astype(float)

# ── 7. INCOME SOURCE dummies (col 8) ─────────────────────────────────────────
income_src_key = {
    "Salary":                       "Income_Salary",
    "Student allowance":            "Income_StudentAllowance",
    "Family support":               "Income_FamilySupport",
    "Freelance/self-employment":    "Income_Freelance",
    "Student job":                  "Income_StudentJob",
    "Government support/benefits":  "Income_GovernmentBenefits",
    "Volunteer work":               "Income_VolunteerWork",
    "Nothing":                      "Income_Nothing",
}
income_dummy_cols = list(income_src_key.values())
income_dummies = pd.DataFrame(0, index=range(n), columns=income_dummy_cols)
inc_src_raw = raw.iloc[:, 8].astype(str)
inc_src_unmatched = set()
for i, cell in enumerate(inc_src_raw):
    if cell.strip().lower() == "nan":
        continue
    for part in [p.strip() for p in cell.split(",")]:
        if part == "":
            continue
        if part in income_src_key:
            income_dummies.at[i, income_src_key[part]] = 1
        else:
            inc_src_unmatched.add(part)
if inc_src_unmatched:
    unmatched["Income_source (unmapped tokens)"] = sorted(inc_src_unmatched)

# ── 8. SHOPPING FREQUENCY (col 9) ────────────────────────────────────────────
shop_freq_map = {
    "Less than once a month":        1,
    "1–2 times per month":           2,
    "3–5 times per month":           3,
    "6–10 times per month":          4,
    "More than 10 times per month":  5,
}
shop_freq_raw = raw.iloc[:, 9].astype(str).str.strip()
shop_freq = track_unmatched("ShoppingFrequency", shop_freq_raw, shop_freq_map)

# ── 9. AVG SPEND PER PURCHASE (col 10) ───────────────────────────────────────
avg_per_map = {
    "€0 – €25":    12,
    "€26 – €50":   38,
    "€51 – €75":   63,
    "€76 – €100":  88,
    "€100+":       125,
}
avg_per_raw = raw.iloc[:, 10].astype(str).str.strip()
avg_per = track_unmatched("AvgPerPurchase", avg_per_raw, avg_per_map)

# ── 10. MONTHLY ONLINE SPEND – Y variable (col 11) ───────────────────────────
monthly_spend_map = {
    "€0 – €50":     25,
    "€51 – €100":   75,
    "€101 – €150": 125,
    "€151 – €200": 175,
    "€200+":       225,
}
ms_raw = raw.iloc[:, 11].astype(str).str.strip()
monthly_spend_Y = track_unmatched("MonthlyOnlineSpend_Y", ms_raw, monthly_spend_map)

# ── 11. BUY category dummies (col 12) ────────────────────────────────────────
buy_key = {
    "Clothing":                                    "Buy_Clothing",
    "Technology/electronics":                      "Buy_Technology",
    "Home items":                                  "Buy_HomeItems",
    "Kitchen items":                               "Buy_KitchenItems",
    "Beauty/personal care":                        "Buy_Beauty",
    "Entertainment (games, subscriptions, etc.)":  "Buy_Entertainment",
    "Merchandise":                                 "Buy_Merchandise",
}
buy_dummy_cols = list(buy_key.values())
buy_dummies = pd.DataFrame(0, index=range(n), columns=buy_dummy_cols)
buy_raw = raw.iloc[:, 12].astype(str)
buy_unmatched_vals = set()

def split_buy_cell(cell):
    """Split on comma, but protect 'Entertainment (games, subscriptions, etc.)'."""
    placeholder = "__ENT__"
    s = cell.replace("Entertainment (games, subscriptions, etc.)", placeholder)
    parts = [p.strip() for p in s.split(",")]
    return [p.replace(placeholder, "Entertainment (games, subscriptions, etc.)") for p in parts]

for i, cell in enumerate(buy_raw):
    if cell.strip().lower() == "nan":
        continue
    for part in split_buy_cell(cell):
        if part == "":
            continue
        if part in buy_key:
            buy_dummies.at[i, buy_key[part]] = 1
        else:
            buy_unmatched_vals.add(part)
if buy_unmatched_vals:
    unmatched["Buy_categories (unmapped tokens)"] = sorted(buy_unmatched_vals)

# ── 12. LIKERT columns (cols 13–37) ──────────────────────────────────────────
likert_map_ci = {
    "strongly disagree": 1,
    "disagree":          2,
    "neutral":           3,
    "agree":             4,
    "strongly agree":    5,
}
likert_names = [
    "WhyShop_Convenient",           # col 13
    "WhyShop_SavesTime",            # col 14
    "WhyShop_BetterPrices",         # col 15
    "WhyShop_EnjoyBrowsing",        # col 16
    "WhyShop_Habit",                # col 17
    "WhyShop_Hobbies",              # col 18
    "WhyShop_Curious",              # col 19
    "WhyShop_Trending",             # col 20
    "Influence_Price",              # col 21
    "Influence_Quality",            # col 22
    "Influence_Brand",              # col 23
    "Influence_SocialMedia",        # col 24
    "Influence_OnlineAds",          # col 25
    "Influence_WebDesign",          # col 26
    "Influence_Reviews",            # col 27
    "Influence_Recommendations",    # col 28
    "Discourages_PoorWebDesign",    # col 29
    "Discourages_NegativeReviews",  # col 30
    "Discourages_HighDeliveryCost", # col 31
    "Discourages_LongDelivery",     # col 32
    "Discourages_PoorReturns",      # col 33
    "Discourages_PaymentSecurity",  # col 34
    "Discourages_LackOfTrust",      # col 35
    "Discourages_LackOfProductInfo",# col 36
    "Discourages_HiddenCosts",      # col 37
]

def map_likert_col(series, col_name):
    result = series.astype(str).str.strip().str.lower().map(likert_map_ci)
    # Identify true mismatches (were not originally NaN/empty)
    orig_notna = series.astype(str).str.strip().str.lower()
    bad = orig_notna[result.isna() & ~orig_notna.isin(["nan", ""])]
    if not bad.empty:
        unmatched[col_name] = bad.unique().tolist()
    return result.astype(float)

likert_series = {}
for idx_offset, name in enumerate(likert_names):
    col_idx = 13 + idx_offset
    likert_series[name] = map_likert_col(raw.iloc[:, col_idx], name)

# ── 13. HOURS ONLINE (col 38) ────────────────────────────────────────────────
hours_map = {
    "Less than 2 hours":  1,
    "2–4 hours":          2,
    "5–7 hours":          3,
    "8–10 hours":         4,
    "More than 10 hours": 5,
}
hours_raw = raw.iloc[:, 38].astype(str).str.strip()
hours_online = track_unmatched("HoursOnline", hours_raw, hours_map)

# ── 14. SOCIAL MEDIA DISCOVERY (col 39) ──────────────────────────────────────
social_map = {
    "Never":           1,
    "Rarely":          2,
    "Sometimes":       3,
    "Often":           4,
    "Very Frequently": 5,
}
social_raw = raw.iloc[:, 39].astype(str).str.strip()
social_discovery = track_unmatched("SocialMediaDiscovery", social_raw, social_map)

# ── 15. ASSEMBLE OUTPUT DATAFRAME ────────────────────────────────────────────
employment_level_map = {
    1: 1,  # Student (non-working) → Not earning
    6: 1,  # Unemployed            → Not earning
    7: 1,  # Stay-at-home          → Not earning
    2: 2,  # Student (working)     → Partially earning
    4: 2,  # Employed part-time    → Partially earning
    3: 3,  # Employed full-time    → Fully earning
    5: 3,  # Self-employed         → Fully earning
}

out = pd.DataFrame()

# Y variable first
out["MonthlyOnlineSpend_Y"] = monthly_spend_Y.values

# Demographics
out["Age"]                = age.values
out["EmploymentLevel"]    = situation.map(employment_level_map).values
out["RelationshipStatus"] = relationship.values
out["MonthlyIncome"]      = monthly_income.values
out["DisposableIncome"]   = disposable.values

# Income source dummies (right after DisposableIncome)
for col in income_dummy_cols:
    out[col] = income_dummies[col].values

# Behavioural shopping variables
out["ShoppingFrequency"] = shop_freq.values
out["AvgPerPurchase"]    = avg_per.values

# Buy dummies (right after AvgPerPurchase)
for col in buy_dummy_cols:
    out[col] = buy_dummies[col].values

# Likert – WhyShop block
why_shop_cols = [
    "WhyShop_Convenient", "WhyShop_SavesTime", "WhyShop_BetterPrices",
    "WhyShop_EnjoyBrowsing", "WhyShop_Habit", "WhyShop_Hobbies",
    "WhyShop_Curious", "WhyShop_Trending",
]
for name in why_shop_cols:
    out[name] = likert_series[name].values
out["WhyShop_Avg"] = out[why_shop_cols].mean(axis=1).round(2)

# Likert – Influence block
influence_cols = [
    "Influence_Price", "Influence_Quality", "Influence_Brand",
    "Influence_SocialMedia", "Influence_OnlineAds", "Influence_WebDesign",
    "Influence_Reviews", "Influence_Recommendations",
]
for name in influence_cols:
    out[name] = likert_series[name].values
out["Influence_Avg"] = out[influence_cols].mean(axis=1).round(2)

# Likert – Discourages block
for name in [
    "Discourages_PoorWebDesign", "Discourages_NegativeReviews",
    "Discourages_HighDeliveryCost", "Discourages_LongDelivery",
    "Discourages_PoorReturns", "Discourages_PaymentSecurity",
    "Discourages_LackOfTrust", "Discourages_LackOfProductInfo",
    "Discourages_HiddenCosts",
]:
    out[name] = likert_series[name].values

# Digital behaviour
out["HoursOnline"]          = hours_online.values
out["SocialMediaDiscovery"] = social_discovery.values

print(f"\nOutput shape: {out.shape}")
print(f"Output columns ({len(out.columns)}):")
for c in out.columns:
    print(f"  {c}")

# ── 16. BUILD CODEBOOK ───────────────────────────────────────────────────────
likert_labels = "1=Strongly Disagree, 2=Disagree, 3=Neutral, 4=Agree, 5=Strongly Agree"

codebook_rows = [
    ("MonthlyOnlineSpend_Y", "Scale (midpoints)",
     "25=€0–€50, 75=€51–€100, 125=€101–€150, 175=€151–€200, 225=€200+"),
    ("Age", "Scale",
     "Numeric integer (years); '25ans ' cleaned to 25"),
    ("EmploymentLevel", "Ordinal",
     "1=Not earning (Student non-working / Unemployed / Stay-at-home), "
     "2=Partially earning (Student working / Employed part-time), "
     "3=Fully earning (Employed full-time / Self-employed). "
     "Collapsed from 7-category employment question into ordered earning capacity levels."),
    ("RelationshipStatus", "Ordinal",
     "1=Single, 2=In a relationship, 3=Cohabiting, 4=Married. "
     "Treated as ordinal by increasing relationship commitment/stability; "
     "messy values normalised before coding."),
    ("MonthlyIncome", "Ordinal",
     "1=€0–€999, 2=€1000–€1999, 3=€2000–€2999, 4=€3000–€3999, 5=€4000+; "
     "NaN=Prefer not to say"),
    ("DisposableIncome", "Ordinal",
     "1=€0–€499, 2=€500–€999, 3=€1000–€1499, 4=€1500+; NaN=Not sure"),
    ("Income_Salary",            "Dummy", "0=No, 1=Yes"),
    ("Income_StudentAllowance",  "Dummy", "0=No, 1=Yes"),
    ("Income_FamilySupport",     "Dummy", "0=No, 1=Yes"),
    ("Income_Freelance",         "Dummy", "0=No, 1=Yes (Freelance/self-employment)"),
    ("Income_StudentJob",        "Dummy", "0=No, 1=Yes"),
    ("Income_GovernmentBenefits","Dummy", "0=No, 1=Yes (Government support/benefits)"),
    ("Income_VolunteerWork",     "Dummy", "0=No, 1=Yes"),
    ("Income_Nothing",           "Dummy", "0=No, 1=Yes"),
    ("ShoppingFrequency", "Ordinal",
     "1=Less than once a month, 2=1–2/month, 3=3–5/month, "
     "4=6–10/month, 5=More than 10/month"),
    ("AvgPerPurchase", "Scale (midpoints)",
     "12=€0–€25, 38=€26–€50, 63=€51–€75, 88=€76–€100, 125=€100+"),
    ("Buy_Clothing",     "Dummy", "0=No, 1=Yes"),
    ("Buy_Technology",   "Dummy", "0=No, 1=Yes (Technology/electronics)"),
    ("Buy_HomeItems",    "Dummy", "0=No, 1=Yes (Home items)"),
    ("Buy_KitchenItems", "Dummy", "0=No, 1=Yes (Kitchen items)"),
    ("Buy_Beauty",       "Dummy", "0=No, 1=Yes (Beauty/personal care)"),
    ("Buy_Entertainment","Dummy", "0=No, 1=Yes (Entertainment: games, subscriptions, etc.)"),
    ("Buy_Merchandise",  "Dummy", "0=No, 1=Yes"),
    ("WhyShop_Convenient",         "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_SavesTime",          "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_BetterPrices",       "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_EnjoyBrowsing",      "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_Habit",              "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_Hobbies",            "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_Curious",            "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_Trending",           "Scale (Likert 1–5)", likert_labels),
    ("WhyShop_Avg",                "Scale (computed)",
     "Row mean of all 8 WhyShop_* columns (1–5). Higher = stronger overall motivation to shop online."),
    ("Influence_Price",            "Scale (Likert 1–5)", likert_labels),
    ("Influence_Quality",          "Scale (Likert 1–5)", likert_labels),
    ("Influence_Brand",            "Scale (Likert 1–5)", likert_labels),
    ("Influence_SocialMedia",      "Scale (Likert 1–5)", likert_labels),
    ("Influence_OnlineAds",        "Scale (Likert 1–5)", likert_labels),
    ("Influence_WebDesign",        "Scale (Likert 1–5)", likert_labels),
    ("Influence_Reviews",          "Scale (Likert 1–5)", likert_labels),
    ("Influence_Recommendations",  "Scale (Likert 1–5)", likert_labels),
    ("Influence_Avg",              "Scale (computed)",
     "Row mean of all 8 Influence_* columns (1–5). Higher = stronger overall influence on purchase decisions."),
    ("Discourages_PoorWebDesign",        "Scale (Likert 1–5)", likert_labels),
    ("Discourages_NegativeReviews",      "Scale (Likert 1–5)", likert_labels),
    ("Discourages_HighDeliveryCost",     "Scale (Likert 1–5)", likert_labels),
    ("Discourages_LongDelivery",         "Scale (Likert 1–5)", likert_labels),
    ("Discourages_PoorReturns",          "Scale (Likert 1–5)", likert_labels),
    ("Discourages_PaymentSecurity",      "Scale (Likert 1–5)", likert_labels),
    ("Discourages_LackOfTrust",          "Scale (Likert 1–5)", likert_labels),
    ("Discourages_LackOfProductInfo",    "Scale (Likert 1–5)", likert_labels),
    ("Discourages_HiddenCosts",          "Scale (Likert 1–5)", likert_labels),
    ("HoursOnline", "Ordinal",
     "1=Less than 2 hours, 2=2–4 hours, 3=5–7 hours, 4=8–10 hours, 5=More than 10 hours"),
    ("SocialMediaDiscovery", "Ordinal",
     "1=Never, 2=Rarely, 3=Sometimes, 4=Often, 5=Very Frequently"),
]

codebook_df = pd.DataFrame(codebook_rows, columns=["ColumnName", "Type", "ValueLabels"])

# ── 17. BUILD REGRESSION SHEET ───────────────────────────────────────────────
regression_cols = [
    "MonthlyOnlineSpend_Y",
    "Age",
    "EmploymentLevel",
    "MonthlyIncome",
    "DisposableIncome",
    "ShoppingFrequency",
    "AvgPerPurchase",
    "WhyShop_Avg",
    "Influence_Avg",
    "HoursOnline",
    "SocialMediaDiscovery",
]
reg_df = out[regression_cols].copy()

# ── 18. WRITE EXCEL ───────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    out.to_excel(writer, sheet_name="Data", index=False)
    codebook_df.to_excel(writer, sheet_name="Codebook", index=False)
    reg_df.to_excel(writer, sheet_name="Regression_Data", index=False)

# ── 18. FORMAT with openpyxl ─────────────────────────────────────────────────
wb = load_workbook(OUTPUT)

# --- Data sheet ---
ws_data = wb["Data"]
hdr_fill   = PatternFill(fill_type="solid", fgColor="1F4E79")   # dark navy
hdr_font   = Font(color="FFFFFF", bold=True, size=10)
alt_fill   = PatternFill(fill_type="solid", fgColor="D9E8F5")   # pale blue
orange_fill = PatternFill(fill_type="solid", fgColor="FF6600")  # orange
red_fill    = PatternFill(fill_type="solid", fgColor="C00000")  # red

# Map column name → 1-based column index for targeted header colouring
col_index = {cell.value: cell.column for cell in ws_data[1]}

for cell in ws_data[1]:
    col_name = cell.value
    if col_name == "WhyShop_Avg":
        cell.fill = orange_fill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
    elif col_name == "Influence_Avg":
        cell.fill = red_fill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
    else:
        cell.fill = hdr_fill
        cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row_idx, row in enumerate(ws_data.iter_rows(min_row=2), start=2):
    if row_idx % 2 == 0:
        for cell in row:
            cell.fill = alt_fill

for col_cells in ws_data.columns:
    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
    ws_data.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 24)

ws_data.freeze_panes = "A2"

# --- Codebook sheet ---
ws_cb = wb["Codebook"]
cb_fill = PatternFill(fill_type="solid", fgColor="375623")  # dark green
cb_font = Font(color="FFFFFF", bold=True, size=10)
cb_alt  = PatternFill(fill_type="solid", fgColor="EAF1E0")  # pale green

for cell in ws_cb[1]:
    cell.fill = cb_fill
    cell.font = cb_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, row in enumerate(ws_cb.iter_rows(min_row=2), start=2):
    if row_idx % 2 == 0:
        for cell in row:
            cell.fill = cb_alt
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws_cb.column_dimensions["A"].width = 32
ws_cb.column_dimensions["B"].width = 20
ws_cb.column_dimensions["C"].width = 85
ws_cb.freeze_panes = "A2"

# --- Key sheet (Likert legend) ---
ws_key = wb.create_sheet("Key")

key_title_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
key_title_font = Font(color="FFFFFF", bold=True, size=11)

# Title row
ws_key["A1"] = "Likert Scale Key (applies to all WhyShop_*, Influence_*, and Discourages_* columns)"
ws_key["A1"].font = key_title_font
ws_key["A1"].fill = key_title_fill
ws_key["A1"].alignment = Alignment(horizontal="left")
ws_key.merge_cells("A1:C1")

# Sub-header
ws_key["A2"] = "Score"
ws_key["B2"] = "Label"
ws_key["C2"] = "Meaning"
sub_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
sub_font = Font(color="FFFFFF", bold=True, size=10)
for col_letter in ("A", "B", "C"):
    ws_key[f"{col_letter}2"].fill = sub_fill
    ws_key[f"{col_letter}2"].font = sub_font
    ws_key[f"{col_letter}2"].alignment = Alignment(horizontal="center")

likert_key_rows = [
    (1, "Strongly Disagree", "Respondent strongly disagrees with the statement"),
    (2, "Disagree",          "Respondent disagrees with the statement"),
    (3, "Neutral",           "Respondent neither agrees nor disagrees"),
    (4, "Agree",             "Respondent agrees with the statement"),
    (5, "Strongly Agree",    "Respondent strongly agrees with the statement"),
]
row_fills = [
    PatternFill(fill_type="solid", fgColor="FDEBD0"),  # light orange-ish for 1
    PatternFill(fill_type="solid", fgColor="FDEBD0"),
    PatternFill(fill_type="solid", fgColor="EAF1E0"),  # neutral green
    PatternFill(fill_type="solid", fgColor="D5E8D4"),
    PatternFill(fill_type="solid", fgColor="D5E8D4"),
]
for i, (score, label, meaning) in enumerate(likert_key_rows, start=3):
    ws_key.cell(row=i, column=1, value=score)
    ws_key.cell(row=i, column=2, value=label)
    ws_key.cell(row=i, column=3, value=meaning)
    for j in range(1, 4):
        ws_key.cell(row=i, column=j).fill = row_fills[i - 3]
        ws_key.cell(row=i, column=j).alignment = Alignment(horizontal="center" if j < 3 else "left")

# Blank row then average column legend
ws_key["A9"]  = "Average Column Key"
ws_key["A9"].font = key_title_font
ws_key["A9"].fill = key_title_fill
ws_key.merge_cells("A9:C9")

ws_key["A10"] = "Column"
ws_key["B10"] = "Colour"
ws_key["C10"] = "Description"
for col_letter in ("A", "B", "C"):
    ws_key[f"{col_letter}10"].fill = sub_fill
    ws_key[f"{col_letter}10"].font = sub_font
    ws_key[f"{col_letter}10"].alignment = Alignment(horizontal="center")

avg_rows = [
    ("WhyShop_Avg",   "Orange", "Mean of 8 'Why do you shop online?' Likert items (1–5)"),
    ("Influence_Avg", "Red",    "Mean of 8 'What influences your purchase decisions?' Likert items (1–5)"),
]
avg_fills = [
    PatternFill(fill_type="solid", fgColor="FFE0CC"),
    PatternFill(fill_type="solid", fgColor="FFCCCC"),
]
for i, (col, colour, desc) in enumerate(avg_rows, start=11):
    ws_key.cell(row=i, column=1, value=col)
    ws_key.cell(row=i, column=2, value=colour)
    ws_key.cell(row=i, column=3, value=desc)
    for j in range(1, 4):
        ws_key.cell(row=i, column=j).fill = avg_fills[i - 11]
        ws_key.cell(row=i, column=j).alignment = Alignment(horizontal="center" if j < 3 else "left")

ws_key.column_dimensions["A"].width = 26
ws_key.column_dimensions["B"].width = 18
ws_key.column_dimensions["C"].width = 65

# --- Regression_Data sheet ---
ws_reg = wb["Regression_Data"]
reg_hdr_fill = PatternFill(fill_type="solid", fgColor="375623")  # dark green
reg_hdr_font = Font(color="FFFFFF", bold=True, size=10)
reg_alt_fill = PatternFill(fill_type="solid", fgColor="EAF1E0")  # pale green

for cell in ws_reg[1]:
    cell.fill = reg_hdr_fill
    cell.font = reg_hdr_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row_idx, row in enumerate(ws_reg.iter_rows(min_row=2), start=2):
    if row_idx % 2 == 0:
        for cell in row:
            cell.fill = reg_alt_fill

for col_cells in ws_reg.columns:
    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
    ws_reg.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 26)

ws_reg.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"\nSaved → {OUTPUT}")

# ── 19. SUMMARY ───────────────────────────────────────────────────────────────
print("\n" + "=" * 62)
print("SUMMARY")
print("=" * 62)
print(f"  Total rows (respondents) : {len(out)}")
print(f"  Total columns in output  : {len(out.columns)}")
dummy_count = sum(1 for c in out.columns if c.startswith(("Income_", "Buy_")))
print(f"    - of which dummies     : {dummy_count}")
print(f"    - of which other vars  : {len(out.columns) - dummy_count}")

print(f"\n  Missing values per column:")
miss = out.isnull().sum()
miss_cols = miss[miss > 0]
if miss_cols.empty:
    print("    (none)")
else:
    for c, v in miss_cols.items():
        print(f"    {c:35s}  {v} missing")

print(f"\n  Unmatched / dirty values (did not fit any mapping):")
if not unmatched:
    print("    (none – all values mapped cleanly)")
else:
    for col, vals in unmatched.items():
        print(f"    [{col}]")
        for v in vals:
            print(f"      '{v}'")
